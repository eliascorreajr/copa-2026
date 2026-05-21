import openpyxl
import json
from datetime import datetime

wb = openpyxl.load_workbook("TABELA-DA-COPA-2026.xlsx", data_only=True)

GROUPS = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]

# Mapping: group letter -> list of 4 teams in order from the spreadsheet
GROUP_TEAMS = {
    "A": ["México", "Coréia do Sul", "República Tcheca", "África do Sul"],
    "B": ["Canadá", "Catar", "Suíça", "Bósnia e Herzegovina"],
    "C": ["Brasil", "Haiti", "Escócia", "Marrocos"],
    "D": ["Estados Unidos", "Austrália", "Turquia", "Paraguai"],
    "E": ["Alemanha", "Costa do Marfim", "Equador", "Curaçao"],
    "F": ["Holanda", "Suécia", "Tunísia", "Japão"],
    "G": ["Bélgica", "Irã", "Nova Zelândia", "Egito"],
    "H": ["Espanha", "Arábia Saudita", "Uruguai", "Cabo Verde"],
    "I": ["França", "Iraque", "Noruega", "Senegal"],
    "J": ["Argentina", "Áustria", "Jordânia", "Argélia"],
    "K": ["Portugal", "Uzbequistão", "Colômbia", "RD do Congo"],
    "L": ["Inglaterra", "Gana", "Panamá", "Croácia"],
}

# Row ranges in the "Grupos" sheet for each pair of groups
# Groups are arranged in pairs: A-B, C-D, E-F, G-H, I-J, K-L
GROUP_ROW_RANGES = [
    ("A", "B", 6, 11),   # rows 6-11 (group header at 5)
    ("C", "D", 14, 19),  # rows 14-19 (header at 13)
    ("E", "F", 22, 27),  # rows 22-27 (header at 21)
    ("G", "H", 30, 35),  # rows 30-35 (header at 29)
    ("I", "J", 38, 43),  # rows 38-43 (header at 37)
    ("K", "L", 46, 51),  # rows 46-51 (header at 45)
]

def parse_datetime_str(dt_str):
    """Parse date string like '11/06/2026 - 16:00' into ISO format."""
    try:
        parts = dt_str.split(" - ")
        date_part = parts[0].strip()
        time_part = parts[1].strip() if len(parts) > 1 else "00:00"
        day, month, year = date_part.split("/")
        hour, minute = time_part.split(":")
        dt = datetime(int(year), int(month), int(day), int(hour), int(minute))
        return dt.isoformat()
    except Exception:
        return dt_str


def extract_group_matches():
    """Extract all 72 group stage matches from the Grupos sheet."""
    ws = wb["Grupos"]
    matches = []
    match_id = 1

    for left_group, right_group, start_row, end_row in GROUP_ROW_RANGES:
        # Determine round number for each row
        rows = list(range(start_row, end_row + 1))
        # Each group has 6 matches, rounds: 1=rows[0,1], 2=rows[2,3], 3=rows[4,5]
        round_map = {}
        for i, row in enumerate(rows):
            round_num = (i // 2) + 1
            round_map[row] = round_num

        for row in rows:
            rnd = round_map[row]

            # Left group match (columns C, D, N)
            date_str = ws.cell(row=row, column=3).value  # Column C
            home_team = ws.cell(row=row, column=4).value   # Column D
            away_team = ws.cell(row=row, column=14).value  # Column N

            if date_str and home_team and away_team:
                iso_date = parse_datetime_str(str(date_str))
                matches.append({
                    "id": match_id,
                    "group": left_group,
                    "round": rnd,
                    "homeTeam": home_team.strip(),
                    "awayTeam": away_team.strip(),
                    "date": iso_date,
                    "status": "notstarted",
                    "homeScore": None,
                    "awayScore": None,
                    "sofaScoreId": None
                })
                match_id += 1

            # Right group match (columns P, Q, AA)
            date_str = ws.cell(row=row, column=16).value  # Column P
            home_team = ws.cell(row=row, column=17).value  # Column Q
            away_team = ws.cell(row=row, column=27).value  # Column AA

            if date_str and home_team and away_team:
                iso_date = parse_datetime_str(str(date_str))
                matches.append({
                    "id": match_id,
                    "group": right_group,
                    "round": rnd,
                    "homeTeam": home_team.strip(),
                    "awayTeam": away_team.strip(),
                    "date": iso_date,
                    "status": "notstarted",
                    "homeScore": None,
                    "awayScore": None,
                    "sofaScoreId": None
                })
                match_id += 1

    return matches


def extract_knockout_matches():
    """Extract all 32 knockout stage matches from the Mata-Mata sheet."""
    ws = wb["Mata-Mata"]

    # Manually extracted from the spreadsheet analysis
    # Each entry: (match_number, stage, date)
    knockout_data = [
        # Fase de 32 (Round of 32) - 16 matches (73-88)
        (73, "round-of-32", datetime(2026, 6, 28)),
        (74, "round-of-32", datetime(2026, 6, 29)),
        (75, "round-of-32", datetime(2026, 6, 29)),
        (76, "round-of-32", datetime(2026, 6, 29)),
        (77, "round-of-32", datetime(2026, 6, 30)),
        (78, "round-of-32", datetime(2026, 6, 30)),
        (79, "round-of-32", datetime(2026, 6, 30)),
        (80, "round-of-32", datetime(2026, 7, 1)),
        (81, "round-of-32", datetime(2026, 7, 1)),
        (82, "round-of-32", datetime(2026, 7, 1)),
        (83, "round-of-32", datetime(2026, 7, 2)),
        (84, "round-of-32", datetime(2026, 7, 2)),
        (85, "round-of-32", datetime(2026, 7, 2)),
        (86, "round-of-32", datetime(2026, 7, 3)),
        (87, "round-of-32", datetime(2026, 7, 3)),
        (88, "round-of-32", datetime(2026, 7, 3)),
        # Oitavas de Final (Round of 16) - 8 matches (89-96)
        (89, "round-of-16", datetime(2026, 7, 4)),
        (90, "round-of-16", datetime(2026, 7, 4)),
        (91, "round-of-16", datetime(2026, 7, 5)),
        (92, "round-of-16", datetime(2026, 7, 5)),
        (93, "round-of-16", datetime(2026, 7, 6)),
        (94, "round-of-16", datetime(2026, 7, 6)),
        (95, "round-of-16", datetime(2026, 7, 7)),
        (96, "round-of-16", datetime(2026, 7, 7)),
        # Quartas de Final (Quarterfinals) - 4 matches (97-100)
        (97, "quarterfinals", datetime(2026, 7, 9)),
        (98, "quarterfinals", datetime(2026, 7, 10)),
        (99, "quarterfinals", datetime(2026, 7, 12)),
        (100, "quarterfinals", datetime(2026, 7, 12)),
        # Semifinais - 2 matches (101-102)
        (101, "semifinals", datetime(2026, 7, 14)),
        (102, "semifinals", datetime(2026, 7, 15)),
        # Disputa de 3º lugar - 1 match (103)
        (103, "third-place", datetime(2026, 7, 18)),
        # Final - 1 match (104)
        (104, "final", datetime(2026, 7, 19)),
    ]

    matches = []
    for match_num, stage, date in knockout_data:
        matches.append({
            "id": match_num,
            "stage": stage,
            "date": date.isoformat(),
            "homeTeam": "TBD",
            "awayTeam": "TBD",
            "homeTeamSeed": None,
            "awayTeamSeed": None,
            "status": "notstarted",
            "homeScore": None,
            "awayScore": None,
            "sofaScoreId": None
        })

    return matches


# Name translation map: Portuguese -> English (for SofaScore API mapping)
NAME_TRANSLATION = {
    "México": "Mexico",
    "Coréia do Sul": "South Korea",
    "República Tcheca": "Czechia",
    "África do Sul": "South Africa",
    "Canadá": "Canada",
    "Catar": "Qatar",
    "Suíça": "Switzerland",
    "Bósnia e Herzegovina": "Bosnia & Herzegovina",
    "Brasil": "Brazil",
    "Haiti": "Haiti",
    "Escócia": "Scotland",
    "Marrocos": "Morocco",
    "Estados Unidos": "USA",
    "Austrália": "Australia",
    "Turquia": "Türkiye",
    "Paraguai": "Paraguay",
    "Alemanha": "Germany",
    "Costa do Marfim": "Côte d'Ivoire",
    "Equador": "Ecuador",
    "Curaçao": "Curaçao",
    "Holanda": "Netherlands",
    "Suécia": "Sweden",
    "Tunísia": "Tunisia",
    "Japão": "Japan",
    "Bélgica": "Belgium",
    "Irã": "Iran",
    "Nova Zelândia": "New Zealand",
    "Egito": "Egypt",
    "Espanha": "Spain",
    "Arábia Saudita": "Saudi Arabia",
    "Uruguai": "Uruguay",
    "Cabo Verde": "Cape Verde",
    "França": "France",
    "Iraque": "Iraq",
    "Noruega": "Norway",
    "Senegal": "Senegal",
    "Argentina": "Argentina",
    "Áustria": "Austria",
    "Jordânia": "Jordan",
    "Argélia": "Algeria",
    "Portugal": "Portugal",
    "Uzbequistão": "Uzbekistan",
    "Colômbia": "Colombia",
    "RD do Congo": "DR Congo",
    "Inglaterra": "England",
    "Gana": "Ghana",
    "Panamá": "Panama",
    "Croácia": "Croatia",
}

SOFASCORE_ROUNDS = {
    1: 1,
    2: 2,
    3: 3,
    "round-of-32": 6,
    "round-of-16": 5,
    "quarterfinals": 27,
    "semifinals": 28,
    "third-place": 50,
    "final": 29,
}


def main():
    group_matches = extract_group_matches()
    knockout_matches = extract_knockout_matches()

    print(f"Group matches: {len(group_matches)}")
    print(f"Knockout matches: {len(knockout_matches)}")
    print(f"Total: {len(group_matches) + len(knockout_matches)}")

    # Validation
    matches_per_group = {}
    for m in group_matches:
        g = m["group"]
        matches_per_group[g] = matches_per_group.get(g, 0) + 1
    print("\nMatches per group:")
    for g in sorted(matches_per_group):
        print(f"  Group {g}: {matches_per_group[g]} matches")

    rounds_count = {}
    for m in group_matches:
        r = m["round"]
        rounds_count[r] = rounds_count.get(r, 0) + 1
    print("\nMatches per round:")
    for r in sorted(rounds_count):
        print(f"  Round {r}: {rounds_count[r]} matches")

    stages_count = {}
    for m in knockout_matches:
        s = m["stage"]
        stages_count[s] = stages_count.get(s, 0) + 1
    print("\nMatches per knockout stage:")
    for s in sorted(stages_count):
        print(f"  {s}: {stages_count[s]} matches")

    # Build final JSON
    data = {
        "tournament": "Copa do Mundo 2026",
        "seasonId": 58210,
        "groups": {g: teams for g, teams in sorted(GROUP_TEAMS.items())},
        "groupMatches": group_matches,
        "knockoutMatches": knockout_matches,
        "nameTranslation": NAME_TRANSLATION,
        "sofaScoreRounds": SOFASCORE_ROUNDS,
        "scoring": {
            "exactScore": 7,
            "correctResult": 3,
            "wrongResult": -1
        },
        "lockTimeMinutes": 30,
        "users": [
            {"email": "admin@bolao.com", "nickname": "AdminSuper", "role": "admin", "hidden": True},
            {"email": "paimei@bolao.com", "nickname": "Pai Mei", "role": "player"},
            {"email": "falso9@bolao.com", "nickname": "Falso 9", "role": "player"},
            {"email": "gildacio@bolao.com", "nickname": "Gildácio", "role": "player"},
            {"email": "whiteglauber@bolao.com", "nickname": "White Glauber", "role": "player"},
        ]
    }

    output_path = "data/matches.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\nJSON saved to {output_path}")
    print(f"Total matches: {len(group_matches) + len(knockout_matches)}")


if __name__ == "__main__":
    main()